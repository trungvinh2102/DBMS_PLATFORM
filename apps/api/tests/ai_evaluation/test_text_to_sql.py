"""
test_text_to_sql.py

Execution accuracy harness for Text-to-SQL evaluation on a deterministic
SQLite sample database.
"""

import argparse
import json
import os
import sqlite3
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import pytest


API_ROOT = Path(__file__).resolve().parents[2]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

CASE_FILE = Path(__file__).with_name("text_to_sql_cases.json")
DEFAULT_REPORT_FILE = Path(__file__).with_name("text_to_sql_results.xlsx")
DEFAULT_SAMPLE_DB_FILE = Path(__file__).with_name("text_to_sql_sample.db")
DEFAULT_SAMPLE_DATABASE_ID = "text-to-sql-sample-sqlite"


@dataclass
class EvalResult:
    id: str
    name: str
    prompt: str
    gold_sql: str
    generated_sql: str
    expected_output: str
    actual_output: str
    status: str
    reason: str
    latency_ms: int


def load_cases() -> list[dict[str, Any]]:
    if not CASE_FILE.exists():
        return []
    return json.loads(CASE_FILE.read_text(encoding="utf-8"))


def filter_cases(
    cases: list[dict[str, Any]],
    case_id: str | None,
    offset: int,
    limit: int | None,
) -> list[dict[str, Any]]:
    if case_id:
        cases = [case for case in cases if case["id"] == case_id]
    elif offset:
        cases = cases[offset:]
    if limit is not None:
        cases = cases[:limit]
    if not cases:
        raise SystemExit("No Text-to-SQL cases matched the requested filters.")
    return cases


def create_sample_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.execute("PRAGMA foreign_keys = ON")
    create_schema(conn)
    seed_data(conn)
    return conn


def create_sample_database_file(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        try:
            path.unlink()
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")
    conn = sqlite3.connect(path)
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        create_schema(conn)
        seed_data(conn)
    finally:
        conn.close()
    return path.resolve()


def connect_sample_database_file(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE departments (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            region TEXT NOT NULL,
            budget INTEGER NOT NULL
        );

        CREATE TABLE employees (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            department_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            salary INTEGER NOT NULL,
            hired_at TEXT NOT NULL,
            is_active INTEGER NOT NULL,
            FOREIGN KEY (department_id) REFERENCES departments(id)
        );

        CREATE TABLE customers (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            city TEXT NOT NULL,
            country TEXT NOT NULL,
            signup_date TEXT NOT NULL,
            status TEXT NOT NULL,
            segment TEXT NOT NULL
        );

        CREATE TABLE categories (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL
        );

        CREATE TABLE suppliers (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            country TEXT NOT NULL
        );

        CREATE TABLE products (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            category_id INTEGER NOT NULL,
            supplier_id INTEGER NOT NULL,
            price REAL NOT NULL,
            cost REAL NOT NULL,
            is_active INTEGER NOT NULL,
            FOREIGN KEY (category_id) REFERENCES categories(id),
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
        );

        CREATE TABLE inventory (
            product_id INTEGER PRIMARY KEY,
            quantity INTEGER NOT NULL,
            reorder_level INTEGER NOT NULL,
            warehouse TEXT NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products(id)
        );

        CREATE TABLE orders (
            id INTEGER PRIMARY KEY,
            customer_id INTEGER NOT NULL,
            employee_id INTEGER NOT NULL,
            order_date TEXT NOT NULL,
            status TEXT NOT NULL,
            shipping_city TEXT NOT NULL,
            discount REAL NOT NULL,
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (employee_id) REFERENCES employees(id)
        );

        CREATE TABLE order_items (
            id INTEGER PRIMARY KEY,
            order_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            unit_price REAL NOT NULL,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        );

        CREATE TABLE payments (
            id INTEGER PRIMARY KEY,
            order_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            payment_method TEXT NOT NULL,
            paid_at TEXT NOT NULL,
            status TEXT NOT NULL,
            FOREIGN KEY (order_id) REFERENCES orders(id)
        );

        CREATE TABLE campaigns (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            channel TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            budget REAL NOT NULL
        );

        CREATE TABLE campaign_orders (
            campaign_id INTEGER NOT NULL,
            order_id INTEGER NOT NULL,
            PRIMARY KEY (campaign_id, order_id),
            FOREIGN KEY (campaign_id) REFERENCES campaigns(id),
            FOREIGN KEY (order_id) REFERENCES orders(id)
        );

        CREATE TABLE support_tickets (
            id INTEGER PRIMARY KEY,
            customer_id INTEGER NOT NULL,
            assigned_employee_id INTEGER,
            subject TEXT NOT NULL,
            priority TEXT NOT NULL,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            resolved_at TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (assigned_employee_id) REFERENCES employees(id)
        );
        """
    )


def seed_data(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        INSERT INTO departments VALUES
        (1,'Sales','North',500000),(2,'Support','South',240000),(3,'Engineering','West',800000),
        (4,'Marketing','East',350000),(5,'Operations','North',300000);

        INSERT INTO employees VALUES
        (1,'Alice Nguyen',1,'Sales Manager',92000,'2020-01-10',1),
        (2,'Ben Carter',1,'Sales Rep',62000,'2021-06-15',1),
        (3,'Cara Lee',2,'Support Lead',70000,'2019-03-20',1),
        (4,'Diego Ruiz',2,'Support Agent',52000,'2022-09-01',1),
        (5,'Emma Smith',3,'Engineer',110000,'2018-11-12',1),
        (6,'Finn Brown',3,'Engineer',98000,'2023-02-18',1),
        (7,'Grace Patel',4,'Marketing Manager',88000,'2020-07-07',1),
        (8,'Henry Kim',5,'Ops Analyst',61000,'2021-12-01',0),
        (9,'Ivy Tran',1,'Sales Rep',64000,'2022-04-22',1),
        (10,'Jack Wilson',4,'Content Specialist',56000,'2023-08-14',1);

        INSERT INTO customers VALUES
        (1,'Acme Corp','buyer@acme.test','New York','USA','2022-01-15','active','enterprise'),
        (2,'Bluebird LLC','ops@bluebird.test','Chicago','USA','2022-03-11','active','smb'),
        (3,'Cedar Retail','team@cedar.test','Toronto','Canada','2021-12-05','inactive','smb'),
        (4,'Delta Foods','it@delta.test','Austin','USA','2023-02-18','active','enterprise'),
        (5,'Evergreen Co','hello@evergreen.test','Seattle','USA','2023-06-21','active','startup'),
        (6,'Futura GmbH','admin@futura.test','Berlin','Germany','2022-10-09','active','enterprise'),
        (7,'Green Mart','contact@greenmart.test','Vancouver','Canada','2023-01-30','active','smb'),
        (8,'Helio Labs','data@helio.test','San Francisco','USA','2024-01-12','trial','startup'),
        (9,'Indigo Stores','buy@indigo.test','London','UK','2021-09-17','active','enterprise'),
        (10,'Juniper Studio','studio@juniper.test','Portland','USA','2024-03-03','trial','startup');

        INSERT INTO categories VALUES
        (1,'Analytics'),(2,'Storage'),(3,'Security'),(4,'Collaboration'),(5,'Training');

        INSERT INTO suppliers VALUES
        (1,'Northwind Supply','USA'),(2,'EuroTech','Germany'),(3,'Maple Source','Canada'),(4,'Pacific Goods','USA');

        INSERT INTO products VALUES
        (1,'Insight Basic',1,1,99,40,1),(2,'Insight Pro',1,1,249,120,1),
        (3,'Vault Lite',2,2,149,80,1),(4,'Vault Enterprise',2,2,599,310,1),
        (5,'Shield Standard',3,3,199,95,1),(6,'Shield Advanced',3,3,399,170,1),
        (7,'Team Hub',4,4,129,60,1),(8,'Team Hub Plus',4,4,229,100,0),
        (9,'SQL Workshop',5,1,499,150,1),(10,'Admin Bootcamp',5,2,799,260,1);

        INSERT INTO inventory VALUES
        (1,120,30,'A'),(2,40,20,'A'),(3,75,25,'B'),(4,10,15,'B'),(5,90,30,'A'),
        (6,18,20,'C'),(7,55,25,'C'),(8,0,10,'A'),(9,12,8,'B'),(10,6,5,'C');

        INSERT INTO orders VALUES
        (1,1,1,'2024-01-05','completed','New York',0.05),(2,2,2,'2024-01-10','completed','Chicago',0),
        (3,4,1,'2024-02-14','completed','Austin',0.1),(4,5,9,'2024-02-20','pending','Seattle',0),
        (5,6,2,'2024-03-02','completed','Berlin',0.05),(6,7,9,'2024-03-18','cancelled','Vancouver',0),
        (7,8,1,'2024-04-04','completed','San Francisco',0.15),(8,9,2,'2024-04-22','completed','London',0),
        (9,1,9,'2024-05-09','refunded','New York',0.2),(10,10,1,'2024-05-20','pending','Portland',0),
        (11,3,2,'2024-06-11','completed','Toronto',0.05),(12,4,9,'2024-06-25','completed','Austin',0),
        (13,6,1,'2024-07-03','completed','Berlin',0.12),(14,2,2,'2024-07-19','completed','Chicago',0),
        (15,5,9,'2024-08-08','completed','Seattle',0.08);

        INSERT INTO order_items VALUES
        (1,1,2,2,249),(2,1,5,1,199),(3,2,1,5,99),(4,2,7,2,129),(5,3,4,1,599),
        (6,3,6,2,399),(7,4,3,1,149),(8,5,10,1,799),(9,5,2,1,249),(10,6,1,3,99),
        (11,7,6,1,399),(12,7,9,1,499),(13,8,4,2,599),(14,8,7,3,129),(15,9,5,2,199),
        (16,10,1,1,99),(17,10,3,1,149),(18,11,9,2,499),(19,12,2,3,249),(20,12,6,1,399),
        (21,13,4,1,599),(22,13,10,1,799),(23,14,7,4,129),(24,14,1,2,99),(25,15,5,1,199),
        (26,15,2,1,249);

        INSERT INTO payments VALUES
        (1,1,662.15,'card','2024-01-06','paid'),(2,2,753,'bank_transfer','2024-01-12','paid'),
        (3,3,1257.3,'card','2024-02-15','paid'),(4,5,995.6,'card','2024-03-03','paid'),
        (5,7,763.3,'paypal','2024-04-05','paid'),(6,8,1585,'bank_transfer','2024-04-24','paid'),
        (7,9,318.4,'card','2024-05-12','refunded'),(8,11,948.1,'card','2024-06-12','paid'),
        (9,12,1146,'paypal','2024-06-26','paid'),(10,13,1230.24,'bank_transfer','2024-07-04','paid'),
        (11,14,714,'card','2024-07-20','paid'),(12,15,412.08,'card','2024-08-09','paid');

        INSERT INTO campaigns VALUES
        (1,'Q1 Enterprise Push','email','2024-01-01','2024-03-31',12000),
        (2,'Spring Startup Trial','social','2024-04-01','2024-05-31',8000),
        (3,'Summer Training Promo','webinar','2024-06-01','2024-08-31',15000);

        INSERT INTO campaign_orders VALUES
        (1,1),(1,3),(1,5),(2,7),(2,10),(3,11),(3,13),(3,15);

        INSERT INTO support_tickets VALUES
        (1,1,3,'Login issue','high','resolved','2024-01-07','2024-01-08'),
        (2,2,4,'Billing question','medium','resolved','2024-01-20','2024-01-22'),
        (3,4,3,'Integration setup','high','open','2024-02-18',NULL),
        (4,5,4,'Trial extension','low','closed','2024-03-01','2024-03-05'),
        (5,6,3,'Security review','high','resolved','2024-03-05','2024-03-06'),
        (6,8,4,'API limit','medium','open','2024-04-07',NULL),
        (7,9,3,'Invoice copy','low','resolved','2024-04-25','2024-04-25'),
        (8,10,NULL,'Onboarding help','medium','open','2024-05-21',NULL),
        (9,1,4,'Refund status','medium','closed','2024-05-30','2024-06-02'),
        (10,3,3,'Migration help','high','resolved','2024-06-15','2024-06-18');
        """
    )
    conn.commit()


def execute_sql(conn: sqlite3.Connection, sql: str) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in conn.execute(sql).fetchall()]


def normalize_rows(rows: list[tuple[Any, ...]]) -> list[tuple[str, ...]]:
    normalized = [tuple("" if value is None else str(value) for value in row) for row in rows]
    return sorted(normalized)


def rows_to_json(rows: list[tuple[Any, ...]]) -> str:
    return json.dumps(rows, ensure_ascii=False, default=str)


def compare_sql(conn: sqlite3.Connection, gold_sql: str, generated_sql: str) -> tuple[bool, str, str, str]:
    try:
        raw_gold_rows = execute_sql(conn, gold_sql)
        gold_rows = normalize_rows(raw_gold_rows)
    except sqlite3.Error as exc:
        return False, f"gold_sql_error: {exc}", "[]", "[]"
    try:
        raw_generated_rows = execute_sql(conn, generated_sql)
        generated_rows = normalize_rows(raw_generated_rows)
    except sqlite3.Error as exc:
        return False, f"generated_sql_error: {exc}", rows_to_json(raw_gold_rows), "[]"
    if gold_rows != generated_rows:
        return False, "execution_result_mismatch", rows_to_json(raw_gold_rows), rows_to_json(raw_generated_rows)
    return True, "execution_result_match", rows_to_json(raw_gold_rows), rows_to_json(raw_generated_rows)


def evaluate_cases(
    cases: list[dict[str, Any]],
    sql_generator: Callable[[dict[str, Any]], str],
    connection_factory: Callable[[], sqlite3.Connection] = create_sample_connection,
) -> list[EvalResult]:
    results = []
    with connection_factory() as conn:
        for case in cases:
            started = time.perf_counter()
            generated_sql = sql_generator(case)
            is_correct, reason, expected_output, actual_output = compare_sql(conn, case["gold_sql"], generated_sql)
            results.append(
                EvalResult(
                    id=case["id"],
                    name=case["name"],
                    prompt=case["prompt"],
                    gold_sql=case["gold_sql"],
                    generated_sql=generated_sql,
                    expected_output=expected_output,
                    actual_output=actual_output,
                    status="PASS" if is_correct else "FAIL",
                    reason=reason,
                    latency_ms=int((time.perf_counter() - started) * 1000),
                )
            )
    return results


def load_predictions(path: Path) -> dict[str, str]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    return {row["id"]: row["sql"] for row in rows}


def register_sample_database(database_id: str, database_path: Path) -> str:
    from models import Db, SessionLocal

    session = SessionLocal()
    try:
        db = session.query(Db).filter(Db.id == database_id).first()
        config = {"database": str(database_path)}
        if db:
            db.type = "sqlite"
            db.databaseName = "Text2SQL Sample SQLite"
            db.isReadOnly = True
            db.config = config
        else:
            session.add(
                Db(
                    id=database_id,
                    type="sqlite",
                    databaseName="Text2SQL Sample SQLite",
                    isReadOnly=True,
                    config=config,
                )
            )
        session.commit()
        return database_id
    finally:
        session.close()


def ai_service_generator(database_id: str, schema: str, model_id: str | None) -> Callable[[dict[str, Any]], str]:
    from services.ai_service import ai_service

    def generate(case: dict[str, Any]) -> str:
        result = ai_service.generate_sql(
            case["prompt"],
            database_id,
            schema,
            model_id=model_id,
        )
        return result.get("sql") or ""

    return generate


@pytest.mark.parametrize("case", load_cases(), ids=lambda case: case["id"])
def test_gold_sql_executes_on_sample_database(case: dict[str, Any]):
    with create_sample_connection() as conn:
        rows = execute_sql(conn, case["gold_sql"])

    assert isinstance(rows, list)


def test_text_to_sql_gold_baseline_is_100_percent():
    cases = load_cases()
    if not cases:
        pytest.skip(f"{CASE_FILE.name} is required for the Text-to-SQL benchmark")
    results = evaluate_cases(cases, lambda case: case["gold_sql"])

    assert len(results) == 100
    assert all(result.status == "PASS" for result in results)


def print_report(results: list[EvalResult]) -> None:
    passed = sum(1 for result in results if result.status == "PASS")
    total = len(results)
    accuracy = passed / total * 100 if total else 0
    print(f"\nTEXT-TO-SQL EXECUTION ACCURACY: {accuracy:.2f}% ({passed}/{total})")
    for result in results:
        marker = "OK" if result.status == "PASS" else "FAIL"
        print(f"{marker} {result.id} {result.name} [{result.latency_ms}ms] {result.reason}")
        if result.status != "PASS":
            print(f"  prompt: {result.prompt}")
            print(f"  gold: {result.gold_sql}")
            print(f"  generated: {result.generated_sql}")
            print(f"  expected_output: {result.expected_output}")
            print(f"  actual_output: {result.actual_output}")


def writable_report_path(output_path: Path) -> Path:
    if not output_path.exists():
        return output_path
    try:
        with output_path.open("a+b"):
            return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")


def write_excel_report(results: list[EvalResult], output_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path = writable_report_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    details = workbook.create_sheet("Cases")

    passed = sum(1 for result in results if result.status == "PASS")
    total = len(results)
    accuracy = passed / total if total else 0

    summary.append(["Metric", "Value"])
    summary.append(["Total cases", total])
    summary.append(["Passed", passed])
    summary.append(["Failed", total - passed])
    summary.append(["Accuracy", accuracy])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary["B5"].number_format = "0.00%"
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 18

    headers = list(asdict(results[0]).keys()) if results else []
    details.append(headers)
    for cell in details[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    status_fills = {
        "PASS": PatternFill("solid", fgColor="C6EFCE"),
        "FAIL": PatternFill("solid", fgColor="FFC7CE"),
    }
    for result in results:
        details.append(list(asdict(result).values()))
        row_index = details.max_row
        details.cell(row=row_index, column=7).fill = status_fills.get(result.status, PatternFill())

    widths = {
        "A": 14,
        "B": 30,
        "C": 55,
        "D": 70,
        "E": 70,
        "F": 70,
        "G": 70,
        "H": 12,
        "I": 28,
        "J": 12,
    }
    for column, width in widths.items():
        details.column_dimensions[column].width = width
    details.freeze_panes = "A2"
    details.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{details.max_row}"

    workbook.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Text-to-SQL execution accuracy benchmark.")
    parser.add_argument("--mode", choices=["ai", "predictions", "gold"], default="ai")
    parser.add_argument("--predictions", type=Path, help="JSON array with id and sql fields.")
    parser.add_argument("--database-id", default=DEFAULT_SAMPLE_DATABASE_ID, help="QurioDB database id for live AI generation.")
    parser.add_argument("--schema", default="main")
    parser.add_argument("--model-id")
    parser.add_argument("--case-id", help="Run one specific case id, for example T2SQL-001.")
    parser.add_argument("--offset", type=int, default=0, help="Skip the first N cases when running a chunk.")
    parser.add_argument("--limit", type=int, help="Run only the first N matching cases.")
    parser.add_argument("--sample-db", type=Path, default=DEFAULT_SAMPLE_DB_FILE, help="SQLite sample database file used by AI mode.")
    parser.add_argument("--output", type=Path, default=DEFAULT_REPORT_FILE, help="Excel report output path.")
    args = parser.parse_args()

    cases = filter_cases(load_cases(), args.case_id, args.offset, args.limit)
    connection_factory = create_sample_connection
    if args.mode == "gold":
        generator = lambda case: case["gold_sql"]
    elif args.mode == "predictions":
        if not args.predictions:
            raise SystemExit("--predictions is required for predictions mode")
        predictions = load_predictions(args.predictions)
        generator = lambda case: predictions.get(case["id"], "")
    else:
        sample_path = create_sample_database_file(args.sample_db)
        register_sample_database(args.database_id, sample_path)
        os.environ.setdefault("QURIODB_RAG_ENABLED", "true")
        connection_factory = lambda: connect_sample_database_file(sample_path)
        generator = ai_service_generator(args.database_id, args.schema, args.model_id)

    results = evaluate_cases(cases, generator, connection_factory=connection_factory)
    print_report(results)
    output_path = write_excel_report(results, args.output)
    print(f"Excel report written to: {output_path}")


if __name__ == "__main__":
    main()
